from flask import Flask, render_template, request
import openpyxl
import os
from datetime import datetime


app = Flask(__name__)

responses = {
    "hello": ["Hi there! 😊", "How can I assist you today?"],
    "how are you": ["I'm just a bot, but I'm doing great! 😃", "What about you?"],
    "what is your name": ["I'm your friendly Customer Service Bot! 🤖"],
    "bye": ["Goodbye! 👋", "Have a great day!"],
    "default": ["I'm not sure about that. Can you rephrase? 🤔"]
}

EXCEL_FILE = "chat_history.xlsx"


def save_chat_to_excel(user_msg, bot_responses):

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Timestamp", "User Message", "Bot Response"])  
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    last_row = list(ws.iter_rows(values_only=True))[-1] if ws.max_row > 1 else None
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for response in bot_responses:
        if last_row and last_row[1] == user_msg and last_row[2] == response:
            continue
        ws.append([timestamp, user_msg, response])
    wb.save(EXCEL_FILE)


@app.route("/", methods=["GET", "POST"])
def chatbot():
    chat_history = []

    if request.method == "POST":
        user_message = request.form["message"].strip().lower()
        bot_response = responses.get(user_message, responses["default"])

        save_chat_to_excel(user_message, bot_response)

        chat_history.append(("You", user_message))
        for response in bot_response:
            chat_history.append(("Bot", response))

    return render_template("index.html", chat_history=chat_history)


if __name__ == "__main__":
    app.run(debug=True)